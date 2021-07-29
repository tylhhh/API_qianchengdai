"""
openpyxl是读写Excel的Python库，是一个比较综合的工具，能够同时读取和修改Excel文档
openpyxl中有三个不同层次的类，每一个类都有各自的属性和方法：
workbook是一个Excel工作表
worksheet是工作表中的表单
cell是表单中的单元格
row:行
column：列
步骤：
1、导入load_workbook模块
2、打开数据文件，Workbook对象  -- wb = load_workbook("文件路径")
3、根据表单名称选择表单 -- sh = wb["表单名称"]
4、遍历第一行的每一列
5、遍历表单中除去第一行的每一行
6、关闭文件
"""
from openpyxl import load_workbook
# 打开Excel
wb = load_workbook("api_cases.xlsx")

# 获取worksheet
sheets = wb.worksheets
print(sheets)

# 定位表单
sheet_login = wb["登陆"]

# 定位单位格
data = sheet_login.cell(2,5).value  # 下标从1开始
print(data)
print(type(data))  # 输出str
# 将str类型转化为原本的类型dict：eval(data)
print(type(eval(data)))

# 通过active获取当前选中的sheet
sheet = wb.active
print(sheet)

# 获取表单的最大列和最大行，即总行数和总列数
print(sheet_login.max_row)
print(sheet_login.max_column)

# 按行获取表单的所有数据
datas = sheet_login.rows

# 转化成列表，每一行是一个元组，元组里面的成员是cell对象
print(list(sheet_login.rows))
print(list(sheet_login.rows)[0]) # 表单第一列
print(list(sheet_login.rows)[0][0].value) # 表单第一列第一列的值

# 遍历第一行
title = []
for item in list(sheet_login.rows)[0]:
    title.append(item.value)
print(title)

# 遍历表单中除去第一行的每一行
all_datas = []
for row in list(sheet_login.rows)[1:]:
    values = []
    for cell in row:
        values.append(cell.value)
    res = dict(zip(title,values)) # title和每一行数据，打包成字典
    all_datas.append(res)  # 每一条用例打包成字典添加到列表中
print(all_datas)


# 关闭文件
wb.close()
